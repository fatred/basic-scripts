#!/usr/bin/env python

import paramiko
from ciscoconfparse import CiscoConfParse
import time
import os
import argparse
import getpass
import openpyxl

# +++++++++++++++++++++++++++
# setup arg parser
# +++++++++++++++++++++++++++
parser = argparse.ArgumentParser(description='get Int Office switchmap')
parser.add_argument('ip', help='IP of the target switch')
parser.add_argument('--username', help='Username for auth')
parser.add_argument('--password', help='Password to try')
parser.add_argument('--output', help='filename to dump to (default to ip_switchmap.cfg)')
parser.add_argument('--debug', help='Enable Debugging output', action="store_true")

# DEFINE FUNCS


def get_switch_conf(ip, username, password, debug=False):
    """
        Setup the SSH session
        params
            :param ip: (string) IP address to connect to
            :param username: username to login with
            :param password: password for session
            :param debug: (bool) be noisy about what we are doing [default false]
            
        return 
            cisco_conf: CiscoConfParse object

    """
    remote_conn_pre = paramiko.SSHClient()
    remote_conn_pre.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    try:
        remote_conn_pre.connect(ip, username=username, password=password, look_for_keys=False, allow_agent=False)
    except paramiko.AuthenticationException as auth_err:
        print auth_err
        username = raw_input("[*] Enter a new Username to try: ")
        password = getpass.getpass(prompt="[*] Enter a new Password to try: ")
        try:
            remote_conn_pre.connect(ip, username=username, password=password, look_for_keys=False, allow_agent=False)
        except paramiko.AuthenticationException as auth_err:
            print auth_err
            print "[E] that's twice.  UFIXURCREDSUCALLMEYEA?!?"
            exit()
    
    print "[*] SSH connection established to %s" % ip
    
    # Get a vty on the ssh session
    remote_conn = remote_conn_pre.invoke_shell()
    if debug:
        print "[D] Interactive SSH session established"
    
    # debugging the initial connection
    if debug:
        conn_prompt = remote_conn.recv(1000)
        print conn_prompt
        
    # turn off paging in the client (no --- MORE ---) 
    disable_paging(remote_conn)
    if debug: 
        print "[D] Paging disabled"
    
    # Send the show run command to the client
    if debug: 
        print "[D] Pulling base config back!"
    remote_conn.send("\n")
    remote_conn.send("show run\n")
    
    # wait for that to come back
    time.sleep(20)
    
    # read the return buffer into an object
    output = remote_conn.recv(256000)
    
    if debug: 
        print output
    
    # convert object to a string for parsing
    conf_string = str(output).splitlines()
    
    cisco_conf = CiscoConfParse(conf_string, factory=True)
    
    switch_hostname = cisco_conf.find_objects(r"^hostname")[0].text[9:]
    switch_version = cisco_conf.find_objects(r"^version")[0].text[-4:]
    print "[*] Config Parsed: %s | %s" % (switch_hostname, switch_version)
    
    # return that as CiscoConfParse obj
    return cisco_conf


def disable_paging(remote_conn):
    """
    Disables paging of the terminal session to get a single "page" of config output
    :param remote_conn: paramiko object for ssh connection

    """
    remote_conn.send("terminal length 0\n")
    time.sleep(1)
    output = remote_conn.recv(1000)
    return output


def update_in_alist(alist, key, value):
    """
    Provides function to update a value in a list 

    :param alist: list object to search
    :param key: key object to query
    :param value: value to search for

    """
    return [(k, v) if (k != key) else (key, value) for (k, v) in alist]


def update_in_alist_inplace(alist, key, value):
    """
    Provides function to update a value in a list without re-organising the list
    :param alist: list object to search
    :param key: key object to query
    :param value: value to search for

    """
    alist[:] = update_in_alist(alist, key, value)


def dump_conf_to_disk(cisco_conf, name):    
    # dump the settings out to a temp file.
    if cisco_conf.has_line_with('!'):
        # cisco_conf has entries
        cisco_conf.save_as("%s.cfg" % name)
        print "[*] Written requested change script to %s/%s.cfg" % (os.getcwd(), name)
    elif not cisco_conf.has_line_with('!'):
        # cisco_conf is empty
        print "[W] No changes needed"


def add_value_to_cell(sheet, row, column, value):
    # put value into a given cell in a sheet
    did_it_work = sheet.cell(row=row, column=column).value = value
    
    return bool(did_it_work)


def main():    
    # +++++++++++++++++++++++++
    # VARS
    # +++++++++++++++++++++++++
    global output_file
    total_intf_count = 0
    dot1x_intf_count = 0
    static_intf_count = 0
    trunk_intf_count = 0
    
    args = parser.parse_args()
    
    if args.debug:
        debug = True
    elif not args.debug:
        debug = False
    
    #  default credentials. BAD NETADMIN. NO BISCUIT!
    ip = args.ip
    if not args.username:
        username = 'defaultUser'
    else:
        username = args.username
    if not args.password:
        password = getpass.getpass('Password for %s@%s: ' % (username, ip))
    else:
        password = args.password
    if args.output:
        output_file = args.output + '.xlsx'
    else:
        output_file = ip + '_switchport_map.xlsx'
    
    # +++++++++++++++++++++++++
    # Execution 
    # +++++++++++++++++++++++++
    
    # read the target config into the system
    switch_conf = get_switch_conf(ip, username, password)
    
    # make a new sheet
    spreadsheet = openpyxl.Workbook()
    sheet = spreadsheet.active
    sheet.title = 'Port Map'
    # add headings
    sheet.cell(row=1, column=1).value = "interface"
    sheet.cell(row=1, column=2).value = "dot1x|static|trunk"
    sheet.cell(row=1, column=3).value = "port desc"
    sheet.cell(row=1, column=4).value = "access vlan|trunk vlans"
    
    # get all the interface objects
    interfaces = switch_conf.find_objects(r'^interface GigabitEthernet')
    
    # populate the spreadsheet
    for interface in interfaces:
        # increment the total intf count
        total_intf_count += 1
        
        # get the layout correct
        switch = interface.ordinal_list[0]
        port = interface.ordinal_list[2]
        
        # check if its a dot1x port
        if interface.re_search_children(r'dot1x'):
            # increment the dot1x intf count
            dot1x_intf_count += 1
            
            if debug: 
                print "[D] Interface %s is dot1x enabled" % interface.name
            add_value_to_cell(sheet, (54+port if switch == 2 else 2+port), 1, interface.name)
            add_value_to_cell(sheet, (54+port if switch == 2 else 2+port), 2, "dot1x")
            add_value_to_cell(sheet, (54+port if switch == 2 else 2+port), 3, interface.description)
            add_value_to_cell(sheet, (54+port if switch == 2 else 2+port), 4, interface.access_vlan)
            
        else:
            add_value_to_cell(sheet, (54+port if switch == 2 else 2+port), 1, interface.name)
            if interface.access_vlan == 0:
                # increment trunk intf counter
                trunk_intf_count += 1
                
                if debug: 
                    print "[D] Interface %s is trunk port" % interface.name
                add_value_to_cell(sheet, (54+port if switch == 2 else 2+port), 2, "trunk")
                add_value_to_cell(sheet, (54+port if switch == 2 else 2+port), 3, interface.description)
                add_value_to_cell(sheet, (54+port if switch == 2 else 2+port), 4, "all")
            else: 
                # increment access port counter
                static_intf_count += 1
                if debug:
                    print "[D] Interface %s is a static port" % interface.name
                add_value_to_cell(sheet, (54+port if switch == 2 else 2+port), 2, "static")
                add_value_to_cell(sheet, (54+port if switch == 2 else 2+port), 3, interface.description)
                add_value_to_cell(sheet, (54+port if switch == 2 else 2+port), 4, interface.access_vlan)
    
    print "[*] %s interfaces" % total_intf_count
    print "[*]\t %s dot1x interfaces" % dot1x_intf_count
    print "[*]\t %s static interfaces" % static_intf_count
    print "[*]\t %s trunk interfaces" % trunk_intf_count
    
    # save sheet to disk
    spreadsheet.save(output_file)


if __name__ == "__main__":
    main()
